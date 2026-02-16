from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import CreateView, DetailView, ListView
from django.urls import reverse
from django.contrib.auth.mixins import UserPassesTestMixin
from django.contrib.auth.decorators import user_passes_test
from .models import CoilIn, CoilPallet, CoilOut, CoilNumber, Job, SKU
from django.contrib import messages
from .forms import CoilInForm, CoilPalletForm, CoilNumberFormSet, CoilOutForm, SKUForm, JobForm

def has_group(user, group_names):
    if user.is_superuser:
        return True
    if isinstance(group_names, str):
        group_names = [group_names]
    return user.groups.filter(name__in=group_names).exists()

def is_sku_manager(user):
    return has_group(user, 'SKU_Manager')

def is_coil_in(user):
    return has_group(user, 'Coil_In')

def is_coil_out(user):
    return has_group(user, ['Coil_Out', 'Coil_In'])

def is_adjuster(user):
    return has_group(user, 'Adjuster')

def is_viewer(user):
    return has_group(user, ['Viewer', 'Coil_In', 'Coil_Out', 'Adjuster', 'SKU_Manager']) or user.is_authenticated

# Create your views here.
def index(request):
    if request.method == 'POST':
        if not is_sku_manager(request.user):
             messages.error(request, 'คุณไม่มีสิทธิ์เพิ่ม SKU')
             return redirect('coil:index')

        form = SKUForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'บันทึก SKU เรียบร้อยแล้ว')
            return redirect('coil:index')  # Redirect to clear POST data
    else:
        form = SKUForm()
        
    return render(request, 'index.html', {'form': form})

@user_passes_test(is_viewer)
def coilin_list(request):
    coils = CoilIn.objects.all().order_by('-timestamp1')
    return render(request, 'coil/coilin_list.html', {'coils': coils})

class CoilInCreateView(UserPassesTestMixin, CreateView):
    model = CoilIn
    form_class = CoilInForm
    template_name = 'coil/coilin_form.html'

    def test_func(self):
        return is_coil_in(self.request.user)

    def form_valid(self, form):
        # Auto-assign the current user's profile to the user field
        form.instance.user = self.request.user.profile
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('coil:coilin_detail', kwargs={'pk': self.object.pk})

from django.db.models import Sum

class CoilInDetailView(UserPassesTestMixin, DetailView):
    model = CoilIn
    template_name = 'coil/coilin_detail.html'
    context_object_name = 'coil'
    
    def test_func(self):
        return is_viewer(self.request.user)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Calculate total weight of all coils in this lot
        # coil -> coilpallet_set -> coilnumber_set
        total_weight = CoilNumber.objects.filter(coilpallet__coilin=self.object).aggregate(Sum('weight'))['weight__sum']
        context['total_weight'] = total_weight or 0
        return context

@user_passes_test(is_coil_in)
def add_pallet(request, pk, pallet_pk=None):
    """
    Add or edit a pallet.
    - If pallet_pk is None: Create new pallet
    - If pallet_pk is provided: Edit existing pallet
    """
    coil = get_object_or_404(CoilIn, pk=pk)
    pallet = None
    is_edit = False

    if pallet_pk:
        pallet = get_object_or_404(CoilPallet, pk=pallet_pk, coilin=coil)
        is_edit = True

    if request.method == 'POST':
        form = CoilPalletForm(request.POST, instance=pallet)
        formset = CoilNumberFormSet(request.POST, instance=pallet)

        if form.is_valid() and formset.is_valid():
            pallet = form.save(commit=False)
            pallet.coilin = coil
            pallet.save()

            # Save the formset with the pallet as the instance
            instances = formset.save(commit=False)
            for instance in instances:
                instance.coilpallet = pallet
                instance.save()

            # Handle deletions from formset
            for obj in formset.deleted_objects:
                obj.delete()

            return redirect('coil:coilin_detail', pk=coil.pk)
    else:
        form = CoilPalletForm(instance=pallet)
        formset = CoilNumberFormSet(instance=pallet)

    return render(request, 'coil/add_pallet.html', {
        'coil': coil,
        'form': form,
        'formset': formset,
        'is_edit': is_edit,
        'pallet': pallet
    })

@user_passes_test(is_viewer)
def print_labels(request, pk):
    coil = get_object_or_404(CoilIn, pk=pk)
    pallets = coil.coilpallet_set.all()

    # Pre-calculate totals for each pallet
    for pallet in pallets:
        coil_numbers = pallet.coilnumber_set.all()
        pallet.total_weight = sum([c.weight for c in coil_numbers if c.weight])
        pallet.total_count = coil_numbers.count()
        pallet.all_coils = coil_numbers # accessible in template simply

    return render(request, 'coil/print_label.html', {'coil': coil, 'pallets': pallets})

def label_list(request):
    """Display all labels with their coils and pallets"""
    coils = CoilIn.objects.all().order_by('-timestamp2')

    # Pre-calculate pallet info for each coil
    for coil in coils:
        pallets = coil.coilpallet_set.all()
        coil.pallet_count = pallets.count()
        coil.total_coils = sum([p.coilnumber_set.count() for p in pallets])

    return render(request, 'coil/label_list.html', {'coils': coils})

class CoilOutCreateView(UserPassesTestMixin, CreateView):
    model = CoilOut
    form_class = CoilOutForm
    template_name = 'coil/coilout_form.html'

    def test_func(self):
        return is_coil_out(self.request.user)

    def form_valid(self, form):
        # Auto-assign the current user's profile to the user field
        form.instance.user = self.request.user.profile
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('coil:coilout_list')

class CoilOutListView(ListView):
    model = CoilOut
    template_name = 'coil/coilout_list.html'
    context_object_name = 'coilouts'
    ordering = ['-timestamp2']

class CoilOutDetailView(UserPassesTestMixin, DetailView):
    model = CoilOut
    template_name = 'coil/coilout_detail.html'
    context_object_name = 'coilout'
    
    def test_func(self):
        return is_viewer(self.request.user)

# Update and Delete Views

from django.views.generic import UpdateView, DeleteView
from django.urls import reverse_lazy

class CoilInUpdateView(UserPassesTestMixin, UpdateView):
    model = CoilIn
    form_class = CoilInForm
    template_name = 'coil/coilin_form.html'
    
    def test_func(self):
        return is_adjuster(self.request.user)
    
    def get_success_url(self):
        return reverse('coil:coilin_detail', kwargs={'pk': self.object.pk})

class CoilInDeleteView(UserPassesTestMixin, DeleteView):
    model = CoilIn
    template_name = 'coil/confirm_delete.html'
    success_url = reverse_lazy('coil:coilin_list')

    def test_func(self):
        return is_adjuster(self.request.user)

from django.http import JsonResponse
from .models import CoilNumber

@user_passes_test(is_coil_out)
def get_coil_sku(request, pk):
    """API endpoint to get SKU for a coil number (AJAX requests only)"""
    print(f"DEBUG: get_coil_sku called for pk={pk} by user {request.user}")
    try:
        coil_number = CoilNumber.objects.get(pk=pk)
        sku_id = coil_number.coilpallet.type0.id
        # Also return the name for better debugging or if needed
        sku_name = str(coil_number.coilpallet.type0)
        weight = coil_number.weight
        print(f"DEBUG: Found SKU {sku_name} (ID: {sku_id}), Weight: {weight}")
        return JsonResponse({'sku_id': sku_id, 'sku_name': sku_name, 'weight': weight})
    except CoilNumber.DoesNotExist:
        print(f"DEBUG: CoilNumber {pk} not found")
        return JsonResponse({'error': 'Coil not found'}, status=404)
    except Exception as e:
        print(f"DEBUG: Error in get_coil_sku: {e}")
        return JsonResponse({'error': str(e)}, status=500)

@user_passes_test(is_coil_out)
def get_job_details(request, pk):
    """API endpoint to get Job details (name, qty) for a job number (AJAX requests only)"""
    try:
        job = Job.objects.get(pk=pk)
        return JsonResponse({
            'job_name_short': job.job_name_short,
            'job_qty': job.job_qty
        })
    except Job.DoesNotExist:
        return JsonResponse({'error': 'Job not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

class CoilOutUpdateView(UserPassesTestMixin, UpdateView):
    model = CoilOut
    form_class = CoilOutForm
    template_name = 'coil/coilout_form.html'
    
    def test_func(self):
        return is_adjuster(self.request.user)
    
    def get_success_url(self):
        return reverse('coil:coilout_detail', kwargs={'pk': self.object.pk})

class CoilOutDeleteView(UserPassesTestMixin, DeleteView):
    model = CoilOut
    template_name = 'coil/confirm_delete.html'
    success_url = reverse_lazy('coil:coilout_list')
    
    def test_func(self):
        return is_adjuster(self.request.user)

# Job Management Views
class JobListView(UserPassesTestMixin, ListView):
    model = Job
    template_name = 'coil/job_list.html'
    context_object_name = 'jobs'
    ordering = ['-date_job']

    def test_func(self):
        return is_sku_manager(self.request.user)

class JobCreateView(UserPassesTestMixin, CreateView):
    model = Job
    form_class = JobForm
    template_name = 'coil/job_form.html'
    success_url = reverse_lazy('coil:job_list')

    def test_func(self):
        return is_sku_manager(self.request.user)

class JobUpdateView(UserPassesTestMixin, UpdateView):
    model = Job
    form_class = JobForm
    template_name = 'coil/job_form.html'
    success_url = reverse_lazy('coil:job_list')

    def test_func(self):
        return is_sku_manager(self.request.user)

class JobDeleteView(UserPassesTestMixin, DeleteView):
    model = Job
    template_name = 'coil/confirm_delete.html'
    success_url = reverse_lazy('coil:job_list')

    def test_func(self):
        return is_sku_manager(self.request.user)

# List Views for Export
class SKUListView(UserPassesTestMixin, ListView):
    model = SKU
    template_name = 'coil/sku_list.html'
    context_object_name = 'skus'
    ordering = ['Type0', 'Type1']

    def test_func(self):
        return is_viewer(self.request.user)

class CoilPalletListView(UserPassesTestMixin, ListView):
    model = CoilPallet
    template_name = 'coil/coilpallet_list.html'
    context_object_name = 'pallets'
    ordering = ['-coilin__timestamp1']

    def test_func(self):
        return is_viewer(self.request.user)

class CoilNumberListView(UserPassesTestMixin, ListView):
    model = CoilNumber
    template_name = 'coil/coilnumber_list.html'
    context_object_name = 'coilnumbers'
    ordering = ['coilpallet__number', 'number']

    def test_func(self):
        return is_viewer(self.request.user)

# Export Views
from django.http import HttpResponse
import csv
import openpyxl
from openpyxl.styles import Font, Alignment

@user_passes_test(is_viewer)
def export_sku_excel(request):
    """Export SKU data to Excel file"""
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=SKU_Export.xlsx'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'SKU'

    # Define headers
    headers = ['Type0', 'Type1', 'Type2', 'Thickness', 'Width', 'Length',
               'Color', 'Grade', 'Manufacturer', 'Note1', 'Note2']

    # Write headers with styling
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Write data
    skus = SKU.objects.all().select_related('manufacturer')
    for row_num, sku in enumerate(skus, 2):
        worksheet.cell(row=row_num, column=1).value = sku.Type0
        worksheet.cell(row=row_num, column=2).value = sku.Type1
        worksheet.cell(row=row_num, column=3).value = sku.Type2
        worksheet.cell(row=row_num, column=4).value = sku.thickness
        worksheet.cell(row=row_num, column=5).value = sku.width
        worksheet.cell(row=row_num, column=6).value = sku.length
        worksheet.cell(row=row_num, column=7).value = sku.color
        worksheet.cell(row=row_num, column=8).value = sku.grade
        worksheet.cell(row=row_num, column=9).value = str(sku.manufacturer)
        worksheet.cell(row=row_num, column=10).value = sku.note1
        worksheet.cell(row=row_num, column=11).value = sku.note2

    workbook.save(response)
    return response

@user_passes_test(is_viewer)
def export_sku_csv(request):
    """Export SKU data to CSV file (Google Sheets compatible)"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=SKU_Export.csv'

    # Write BOM for Excel compatibility with UTF-8
    response.write('\ufeff'.encode('utf8'))

    writer = csv.writer(response)

    # Write headers
    writer.writerow(['Type0', 'Type1', 'Type2', 'Thickness', 'Width', 'Length',
                     'Color', 'Grade', 'Manufacturer', 'Note1', 'Note2'])

    # Write data
    skus = SKU.objects.all().select_related('manufacturer')
    for sku in skus:
        writer.writerow([
            sku.Type0, sku.Type1, sku.Type2, sku.thickness, sku.width,
            sku.length, sku.color, sku.grade, str(sku.manufacturer),
            sku.note1, sku.note2
        ])

    return response

@user_passes_test(is_viewer)
def export_coilpallet_excel(request):
    """Export CoilPallet data to Excel file"""
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=CoilPallet_Export.xlsx'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'CoilPallet'

    # Define headers
    headers = ['Pallet Number', 'Lot', 'SKU', 'Supplier', 'Owner', 'Timestamp']

    # Write headers with styling
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Write data
    pallets = CoilPallet.objects.all().select_related('coilin', 'type0', 'coilin__supplier', 'coilin__owner')
    for row_num, pallet in enumerate(pallets, 2):
        worksheet.cell(row=row_num, column=1).value = pallet.number
        worksheet.cell(row=row_num, column=2).value = pallet.coilin.lot if pallet.coilin else ''
        worksheet.cell(row=row_num, column=3).value = str(pallet.type0)
        worksheet.cell(row=row_num, column=4).value = str(pallet.coilin.supplier) if pallet.coilin and pallet.coilin.supplier else ''
        worksheet.cell(row=row_num, column=5).value = str(pallet.coilin.owner) if pallet.coilin and pallet.coilin.owner else ''
        worksheet.cell(row=row_num, column=6).value = str(pallet.coilin.timestamp1) if pallet.coilin and pallet.coilin.timestamp1 else ''

    workbook.save(response)
    return response

@user_passes_test(is_viewer)
def export_coilpallet_csv(request):
    """Export CoilPallet data to CSV file (Google Sheets compatible)"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=CoilPallet_Export.csv'

    # Write BOM for Excel compatibility with UTF-8
    response.write('\ufeff'.encode('utf8'))

    writer = csv.writer(response)

    # Write headers
    writer.writerow(['Pallet Number', 'Lot', 'SKU', 'Supplier', 'Owner', 'Timestamp'])

    # Write data
    pallets = CoilPallet.objects.all().select_related('coilin', 'type0', 'coilin__supplier', 'coilin__owner')
    for pallet in pallets:
        writer.writerow([
            pallet.number,
            pallet.coilin.lot if pallet.coilin else '',
            str(pallet.type0),
            str(pallet.coilin.supplier) if pallet.coilin and pallet.coilin.supplier else '',
            str(pallet.coilin.owner) if pallet.coilin and pallet.coilin.owner else '',
            str(pallet.coilin.timestamp1) if pallet.coilin and pallet.coilin.timestamp1 else ''
        ])

    return response

@user_passes_test(is_viewer)
def export_coilnumber_excel(request):
    """Export CoilNumber data to Excel file"""
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=CoilNumber_Export.xlsx'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'CoilNumber'

    # Define headers
    headers = ['Coil Number', 'Weight (kg)', 'Pallet Number', 'Lot', 'SKU']

    # Write headers with styling
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Write data
    coils = CoilNumber.objects.all().select_related('coilpallet', 'coilpallet__coilin', 'coilpallet__type0')
    for row_num, coil in enumerate(coils, 2):
        worksheet.cell(row=row_num, column=1).value = coil.number
        worksheet.cell(row=row_num, column=2).value = coil.weight
        worksheet.cell(row=row_num, column=3).value = coil.coilpallet.number if coil.coilpallet else ''
        worksheet.cell(row=row_num, column=4).value = coil.coilpallet.coilin.lot if coil.coilpallet and coil.coilpallet.coilin else ''
        worksheet.cell(row=row_num, column=5).value = str(coil.coilpallet.type0) if coil.coilpallet and coil.coilpallet.type0 else ''

    workbook.save(response)
    return response

@user_passes_test(is_viewer)
def export_coilnumber_csv(request):
    """Export CoilNumber data to CSV file (Google Sheets compatible)"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=CoilNumber_Export.csv'

    # Write BOM for Excel compatibility with UTF-8
    response.write('\ufeff'.encode('utf8'))

    writer = csv.writer(response)

    # Write headers
    writer.writerow(['Coil Number', 'Weight (kg)', 'Pallet Number', 'Lot', 'SKU'])

    # Write data
    coils = CoilNumber.objects.all().select_related('coilpallet', 'coilpallet__coilin', 'coilpallet__type0')
    for coil in coils:
        writer.writerow([
            coil.number,
            coil.weight,
            coil.coilpallet.number if coil.coilpallet else '',
            coil.coilpallet.coilin.lot if coil.coilpallet and coil.coilpallet.coilin else '',
            str(coil.coilpallet.type0) if coil.coilpallet and coil.coilpallet.type0 else ''
        ])

    return response
